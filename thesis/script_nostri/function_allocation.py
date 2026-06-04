from __future__ import annotations

from parameters import *

import pandas as pd
import geopandas as gpd
from pathlib import Path
import matplotlib.pyplot as plt 
import numpy as np
import rasterio
from rasterio.mask import mask as raster_mask
import shapely
from shapely.geometry import Point, box, shape, LineString
from skimage.graph import MCP_Geometric
from pyproj import Geod, CRS
from onstove.layer import VectorLayer, RasterLayer
from shapely.ops import nearest_points
from scipy.optimize import linprog
from openpyxl import Workbook
import warnings
import contextlib
import heapq
import io
import json
import math
import os
import shutil
import sqlite3
import time
from pathlib import Path
from typing import Dict, List, Sequence, Tuple, Optional
import rioxarray
import xarray as xr
from onstove import OnStove
from rasterio.features import geometry_mask, shapes
from rasterio.transform import rowcol
from rasterio.warp import Resampling, reproject, transform
from scipy.sparse import csr_matrix
from scipy.sparse.csgraph import connected_components
from scipy.spatial import cKDTree
import function_tools as tool
import function_costs as cost
import function_process as process
from dataclasses import dataclass

def build_surface(friction_path, ref_profile=None):
    """
    Build a surface from friction data.
    If ref_profile is provided, reprojects the friction raster to match the reference grid.
    Returns the cost array, transform, geodesic helper, and CRS.
    """
    if ref_profile is not None:
        # Reproject on the fly to match the master reference (e.g., EPSG:3857)
        cost_layer = tool.reproject_to_reference(friction_path, ref_profile, Resampling.nearest)
        
        # Handle nodata
        cost_layer[np.isnan(cost_layer)] = np.inf
        cost_layer[cost_layer < 0] = np.inf
        
        transform = ref_profile['transform']
        crs_obj = ref_profile['crs']
    else:
        # Fallback to original logic if no reference is provided
        friction = RasterLayer(name='friction_surface', path=str(friction_path))
        cost_layer = friction.data.astype(np.float64).copy()
        
        nodata = friction.meta.get('nodata')
        cost_layer[np.isnan(cost_layer)] = np.inf
        if nodata is not None:
            cost_layer[cost_layer == nodata] = np.inf
        cost_layer[cost_layer < 0] = np.inf
        
        transform = friction.meta['transform']
        crs_obj = friction.meta.get('crs')
    
    use_geodesic = bool(getattr(crs_obj, 'is_geographic', False))
    geod = Geod(ellps='WGS84') if use_geodesic else None
    
    return cost_layer, transform, geod, crs_obj


def calculate_mcp_routes(origin_row, origin_col, dest_rows, dest_cols, cost_layer, transform, geod):
    """
    Compute travel time and path distance from one origin to many destinations.
    Uses MCP to find least-time paths on the cost surface.
    Returns arrays of minutes and km aligned to destination lists.
    Invalid destinations remain NaN.
    """
    max_row, max_col = cost_layer.shape
    times_min = np.full(len(dest_rows), np.nan, dtype=np.float64)
    distances_km = np.full(len(dest_rows), np.nan, dtype=np.float64)

    if not (0 <= origin_row < max_row and 0 <= origin_col < max_col):
        return times_min, distances_km

    mcp = MCP_Geometric(cost_layer, fully_connected=True)
    cumulative_costs, _ = mcp.find_costs(starts=np.array([[origin_row, origin_col]], dtype=np.int64))

    for idx, (d_row, d_col) in enumerate(zip(dest_rows, dest_cols)):
        if 0 <= d_row < max_row and 0 <= d_col < max_col:
            val = cumulative_costs[d_row, d_col]
            if np.isfinite(val):
                times_min[idx] = float(val)
                
                path = mcp.traceback((d_row, d_col))
                if len(path) <= 1:
                    distances_km[idx] = 0.0
                    continue
                    
                p_rows = np.array([p[0] for p in path], dtype=np.int64)
                p_cols = np.array([p[1] for p in path], dtype=np.int64)
                xs, ys = rasterio.transform.xy(transform, p_rows, p_cols, offset='center')
                
                xs = np.asarray(xs, dtype=np.float64)
                ys = np.asarray(ys, dtype=np.float64)

                if geod is not None:
                    _, _, seg_m = geod.inv(xs[:-1], ys[:-1], xs[1:], ys[1:])
                    distances_km[idx] = float(np.nansum(np.abs(seg_m)) / 1000.0)
                else:
                    dx = np.diff(xs)
                    dy = np.diff(ys)
                    distances_km[idx] = float(np.nansum(np.hypot(dx, dy)) / 1000.0)
                    
    return times_min, distances_km

def assign_nearest(
    layer1: gpd.GeoDataFrame,
    layer2: gpd.GeoDataFrame,
    layer1_name: str,
    layer1_id_col: str,
    surface_data: tuple = None,
    friction_path: str = None
) -> gpd.GeoDataFrame:
    """
    For each feature in layer2, find the nearest feature in layer1 by least‑time path.
    
    Adds three columns to layer2:
        nearest_{layer1_name}_id       – ID of the nearest feature in layer1
        nearest_{layer1_name}_distance – One‑way path distance (km)
        nearest_{layer1_name}_time     – One‑way travel time (minutes)
    
    Parameters
    ----------
    layer1 : gpd.GeoDataFrame
        Origin points (must have point geometries and an ID column).
    layer2 : gpd.GeoDataFrame
        Destination points (point geometries).
    layer1_name : str
        Name of the origin layer, used as suffix for output columns (e.g., 'primary_storage').
    surface_data : tuple, optional
        Precomputed cost surface (cost_layer, transform, geod, crs_obj).
    friction_path : str, optional
        Path to friction raster; required if surface_data is None.
    layer1_id_col : str, default 'id_supply'
        Column in layer1 that uniquely identifies each feature.
    
    Returns
    -------
    gpd.GeoDataFrame
        A copy of layer2 with the three additional columns.
    """
    result = layer2.copy()
    
    if result.empty:
        for suffix in ['_id', '_distance', '_time']:
            result[f"nearest_{layer1_name}_{suffix}"] = None
        return result
    
    if layer1.empty:
        raise ValueError("Layer1 (origin) GeoDataFrame is empty; cannot assign.")
    
    # Build or retrieve cost surface
    if surface_data is None:
        if friction_path is None:
            raise ValueError("Either surface_data or friction_path must be provided")
        surface_data = build_surface(friction_path)   # ensure function exists
    cost_layer, transform, geod, crs_obj = surface_data
    
    # Reproject both layers to the raster CRS
    layer1_proj = layer1.to_crs(crs_obj)
    layer2_proj = result.to_crs(crs_obj)
    
    # Get pixel coordinates
    def get_pixel_coords(gdf):
        rows, cols = [], []
        for geom in gdf.geometry:
            if geom.is_empty:
                rows.append(-1)
                cols.append(-1)
            else:
                r, c = rowcol(transform, geom.x, geom.y)
                rows.append(r)
                cols.append(c)
        return np.array(rows), np.array(cols)
    
    o_rows, o_cols = get_pixel_coords(layer1_proj)
    d_rows, d_cols = get_pixel_coords(layer2_proj)
    
    n_origins = len(layer1)
    n_dests = len(result)
    
    best_origin_idx = np.full(n_dests, -1, dtype=int)
    best_times = np.full(n_dests, np.inf, dtype=float)
    best_dists = np.full(n_dests, np.nan, dtype=float)
    
    # For each origin, compute routes to all destinations and update best
    for i in range(n_origins):
        if o_rows[i] < 0 or o_cols[i] < 0:
            continue
        times_min, dists = calculate_mcp_routes(
            o_rows[i], o_cols[i],
            d_rows, d_cols,
            cost_layer, transform, geod
        )
        # times_min in minutes, dists in km
        for j in range(n_dests):
            t = times_min[j]
            if np.isfinite(t) and t < best_times[j]:
                best_times[j] = t
                best_dists[j] = dists[j]
                best_origin_idx[j] = i
    
    # Prepare output columns
    id_col = f"nearest_{layer1_name}_id"
    dist_col = f"nearest_{layer1_name}_distance"
    time_col = f"nearest_{layer1_name}_time"
    
    result[id_col] = None
    result[dist_col] = np.nan
    result[time_col] = np.nan
    
    origin_ids = layer1[layer1_id_col].astype(str).to_numpy()
    
    for j in range(n_dests):
        if best_origin_idx[j] >= 0:
            result.at[result.index[j], id_col] = origin_ids[best_origin_idx[j]]
            result.at[result.index[j], dist_col] = best_dists[j]
            result.at[result.index[j], time_col] = best_times[j]
    
    return result

import pandas as pd
import numpy as np
import geopandas as gpd

def calculate_percentages_supply(
    refineries_gdf, ports_gdf, gas_plants_gdf, border_points_gdf, nat_shares_df,
    refineries_cols='crude_capacity', 
    ports_cols='LPG_capacity', 
    gas_plants_cols='LPG_prod', 
    border_points_cols=None
    ):
    """
    Calculate facility-level market shares from national shares and capacities.
    Accepts single strings or lists of strings for the weight columns.
    Averages the calculated percentages across multiple columns and rescales.
    """
    target_crs = refineries_gdf.crs or ports_gdf.crs or gas_plants_gdf.crs or border_points_gdf.crs

    specs = {
        'refineries': (refineries_gdf, refineries_cols),
        'ports': (ports_gdf, ports_cols),
        'gas_plants': (gas_plants_gdf, gas_plants_cols),
        'border_points': (border_points_gdf, border_points_cols) 
    }

    by_category = {}

    for category, (src_gdf, weight_cols) in specs.items():
        gdf = src_gdf.copy()

        if target_crs and gdf.crs and gdf.crs != target_crs:
            gdf = gdf.to_crs(target_crs)

        if 'name' not in gdf.columns:
            gdf['name'] = f"{category}_" + gdf.index.astype(str)
            
        if 'cost_source' in gdf.columns:
            gdf['cost_source'] = pd.to_numeric(gdf['cost_source'], errors='coerce')
        else:
            gdf['cost_source'] = np.nan
        
        if 'country' in gdf.columns:
            gdf['country'] = gdf['country'].fillna('Unknown')
        else:
            gdf['country'] = 'Unknown'

        gdf['category'] = 'gasplants' if category == 'gas_plants' else category

        if weight_cols:
            # Ensure weight_cols is a list for iteration
            if isinstance(weight_cols, str):
                weight_cols = [weight_cols]

            avg_shares = pd.Series(0.0, index=gdf.index, dtype='float64')
            valid_cols_count = 0

            for col in weight_cols:
                if col in gdf.columns:
                    weights = pd.to_numeric(gdf[col], errors='coerce').fillna(0.0)
                    total_weight = weights.sum()
                    
                    if total_weight > 0:
                        # Calculate the fractional share for this specific column and add to total
                        avg_shares += (weights / total_weight)
                        valid_cols_count += 1

            # Average the shares across all provided valid columns
            if valid_cols_count > 0:
                avg_shares = avg_shares / valid_cols_count

            # Scale by the national share
            nat_share = nat_shares_df.loc[category].iloc[0] / 100.0
            gdf['percentage'] = avg_shares * nat_share
            
        else:
            if 'percentage' in gdf.columns:
                gdf['percentage'] = pd.to_numeric(gdf['percentage'], errors='coerce').fillna(0.0)
            else:
                gdf['percentage'] = 0.0

        by_category[category] = gdf

    result = gpd.GeoDataFrame(pd.concat(list(by_category.values()), ignore_index=True), crs=target_crs)

    print(f"Check: {len(result)} sources calculated | Total share: {result['percentage'].sum():.2%}")
    return result, by_category


def calculate_percentages_filling(filling_gdf, filling_cols=None):
    """
    Calculate demand-based percentage shares for filling points using one or multiple columns.
    Averages the proportional shares across columns and normalizes the final result.
    """
    filling = filling_gdf.copy()

    if filling_cols is None:
        priority_cols = ["percentage_demand", "total_fil_clients", "clients"]
        found_col = next((c for c in priority_cols if c in filling.columns), None)
        if not found_col:
            raise ValueError(
                "No valid demand-related column found. Looked for: "
                "percentage_demand, total_fil_clients, or clients"
            )
        filling_cols = [found_col]

    # Ensure filling_cols is a list
    if isinstance(filling_cols, str):
        filling_cols = [filling_cols]

    missing_cols = [c for c in filling_cols if c not in filling.columns]
    if missing_cols:
        raise ValueError(f"Columns not found in GeoDataFrame: {missing_cols}")

    avg_shares = pd.Series(0.0, index=filling.index, dtype='float64')
    valid_cols_count = 0

    for col in filling_cols:
        demand_vals = pd.to_numeric(filling[col], errors="coerce").fillna(0.0)
        total_demand = demand_vals.sum()

        if total_demand > 0:
            if col == "percentage_demand":
                # Assume percentage_demand is already a 0-1 fraction
                avg_shares += demand_vals
            else:
                # Convert raw counts to a 0-1 fraction
                avg_shares += (demand_vals / total_demand)
            valid_cols_count += 1

    if valid_cols_count > 0:
        # Average the fractions and apply a final normalization to ensure the sum is exactly 1.0
        averaged = avg_shares / valid_cols_count
        final_total = averaged.sum()
        
        if final_total > 0:
            filling["percentage_demand"] = averaged / final_total
        else:
            filling["percentage_demand"] = 0.0
    else:
        filling["percentage_demand"] = 0.0

    print(f"Check: Demand shares calculated using columns: {filling_cols}")
    return filling

def calculate_percentages(
    step, 
    refineries_gdf=None, 
    ports_gdf=None, 
    gas_plants_gdf=None, 
    border_points_gdf=None, 
    filling_points_gdf=None, 
    nat_shares_df=None,
    **kwargs
    ):
    """
    Wrapper function to calculate percentages for both supply and filling stages.
    Accepts optional column names/dictionaries via **kwargs to override default weight columns.
    
    Optional kwargs for 'supply': refineries_cols, ports_cols, gas_plants_cols, border_points_cols
    Optional kwargs for 'filling': filling_cols
    """
    if step == "supply":
        # Filter kwargs to only include those accepted by the updated supply function
        supply_kwargs = {
            k: v for k, v in kwargs.items() 
            if k in ['refineries_cols', 'ports_cols', 'gas_plants_cols', 'border_points_cols']
        }
        return calculate_percentages_supply(
            refineries_gdf, ports_gdf, gas_plants_gdf, border_points_gdf, nat_shares_df, **supply_kwargs
        )
        
    elif step == "filling":
        # Filter kwargs to only include those accepted by the updated filling function
        filling_kwargs = {
            k: v for k, v in kwargs.items() 
            if k == 'filling_cols'
        }
        return calculate_percentages_filling(filling_points_gdf, **filling_kwargs)
        
    else:
        raise ValueError("Invalid step specified. Use 'supply' or 'filling'.")


def aggregate_supply_to_storage(primary_storage_gdf, refineries_gdf, ports_gdf, gas_plants_gdf, border_points_gdf): 
    """
    Aggregate supply percentages from all sources (refineries, ports, gas plants, border points)
    to primary storage facilities at the storage level.
    """
    ps = primary_storage_gdf.copy()
    ps["id_supply"] = ps["id_supply"].astype(str)
    
    # Collect all supply layers
    supply_layers = [refineries_gdf, ports_gdf, gas_plants_gdf, border_points_gdf]
    supply_layers = [g for g in supply_layers if g is not None and not g.empty]
    
    if supply_layers:
        # Align CRS across all layers
        target_crs = next((g.crs for g in supply_layers if g.crs is not None), None)
        if target_crs is not None:
            aligned_layers = [g.to_crs(target_crs) if g.crs is not None and g.crs != target_crs else g for g in supply_layers]
        else:
            aligned_layers = supply_layers
        
        all_supply = pd.concat(aligned_layers, ignore_index=True)
    else:
        all_supply = gpd.GeoDataFrame()
    
    # Aggregate supply by tank id (nearest_primary_storage_id)
    if not all_supply.empty and "nearest_primary_storage_id" in all_supply.columns:
        supply_sum = all_supply.groupby("nearest_primary_storage_id")["percentage"].sum()
    else:
        supply_sum = pd.Series(dtype="float64")
    
    ps["percentage_supply"] = ps["id_supply"].map(supply_sum).fillna(0.0)
    
    print(f"✓ Supply aggregated to {len(ps)} storage facilities | Total supply: {ps['percentage_supply'].sum():.4f}")
    
    return ps


def aggregate_filling_to_storage(primary_storage_gdf, filling_points_gdf): 
    """
    Aggregate filling percentages from all filling points to primary storage facilities
    at the storage level.
    """
    ps = primary_storage_gdf.copy()
    ps["id_supply"] = ps["id_supply"].astype(str)
    
    # Aggregate filling percentages by nearest_primary_storage_id
    if filling_points_gdf is not None and not filling_points_gdf.empty and "nearest_primary_storage_id" in filling_points_gdf.columns:
        filling_sum = filling_points_gdf.groupby("nearest_primary_storage_id")["percentage_demand"].sum()
    else:
        filling_sum = pd.Series(dtype="float64")
    
    ps["percentage_demand"] = ps["id_supply"].map(filling_sum).fillna(0.0)
    
    print(f"✓ Filling aggregated to {len(ps)} storage facilities | Total filling: {ps['percentage_demand'].sum():.4f}")

    return ps

def aggregate_to_storage(stage: str,
                        primary_storage_gdf, 
                         refineries_gdf = None, 
                         ports_gdf = None, 
                         gas_plants_gdf = None, 
                         border_points_gdf = None, 
                         filling_points_gdf = None):
    """
    apply aggregation to storage for both supply and filling, depending on the stage of the process.
    """
    if stage == "supply":
        ps = aggregate_supply_to_storage(primary_storage_gdf, refineries_gdf, ports_gdf, gas_plants_gdf, border_points_gdf)
    elif stage == "filling":
        ps = aggregate_filling_to_storage(primary_storage_gdf, filling_points_gdf)
    else:
        raise ValueError("Invalid stage specified. Use 'supply' or 'filling'.")
    
    return ps


def build_storage_routing_matrices(primary_storage_gdf, friction_path):
    """
    Build travel time and distance matrices for inter-storage routing.
    Computes N×N matrices where N = number of storage facilities.
    """
    ps = primary_storage_gdf.copy()
    n_storage = len(ps)
    storage_ids = ps.index.tolist()
    
    # Build cost surface
    cost_layer, transform, geod, crs_obj = build_surface(friction_path)
    
    if crs_obj is None:
        raise ValueError("Friction raster has no CRS in metadata")
    
    # Project storage to friction CRS
    storage_tt = ps.to_crs(crs_obj).copy()
    storage_points = storage_tt.geometry
    if not storage_tt.geom_type.eq('Point').all():
        storage_points = storage_tt.geometry.centroid
    
    # Precompute pixel coordinates
    storage_rows, storage_cols = [], []
    for geom in storage_points:
        r, c = rasterio.transform.rowcol(transform, geom.x, geom.y)
        storage_rows.append(r)
        storage_cols.append(c)
    
    # Compute N×N routing matrices
    travel_time_matrix = np.full((n_storage, n_storage), np.inf, dtype=np.float64)
    distance_matrix = np.full((n_storage, n_storage), np.nan, dtype=np.float64)
    
    for i, (orig_row, orig_col) in enumerate(zip(storage_rows, storage_cols)):
        t_times, t_dists = calculate_mcp_routes(
            orig_row, orig_col, storage_rows, storage_cols, cost_layer, transform, geod
        )
        travel_time_matrix[i, :] = t_times
        distance_matrix[i, :] = t_dists
    
    print(f"✓ Routing matrices computed for {n_storage} storage facilities\n")
    return travel_time_matrix, distance_matrix, crs_obj, storage_ids


def optimize_storage_allocation(primary_storage_gdf, travel_time_matrix):
    """
    Solve Linear Programming problem to optimize inter-storage allocation.
    Minimizes total travel time subject to supply/demand balance constraints.
    """
    ps = primary_storage_gdf.copy()
    n_storage = len(ps)
    
    # Flatten cost vector (travel times)
    c = travel_time_matrix.flatten()
    
    # Build constraint matrices
    A_eq = []
    b_eq = []
    
    # Supply constraints: sum(x[i,:]) = supply[i]
    for i in range(n_storage):
        row = np.zeros(n_storage * n_storage)
        row[i * n_storage:(i+1) * n_storage] = 1
        A_eq.append(row)
        b_eq.append(float(ps.iloc[i]['percentage_supply']))
    
    # Demand constraints: sum(x[:,j]) = demand[j]
    for j in range(n_storage):
        row = np.zeros(n_storage * n_storage)
        for i in range(n_storage):
            row[i * n_storage + j] = 1
        A_eq.append(row)
        b_eq.append(float(ps.iloc[j]['percentage_demand']))
    
    A_eq = np.array(A_eq)
    b_eq = np.array(b_eq)
    
    #handle case in which some travel times are infinite (no path) by replacing inf with a large finite number
    finite_values = c[np.isfinite(c)]
    max_finite = finite_values.max() if finite_values.size > 0 else 1e6
    c = np.nan_to_num(c, nan=max_finite*10, posinf=max_finite*10)

    # Solve LP
    result = linprog(c, A_eq=A_eq, b_eq=b_eq, bounds=(0, None), method='highs')
    
    if result.success:
        allocation_matrix = result.x.reshape((n_storage, n_storage))
        print(f"✓ Optimization successful\n")
        print(f"  Objective (total travel time): {result.fun:.4f} minutes\n")
    else:
        print(f"✗ Optimization failed: {result.message}\n")
        allocation_matrix = np.zeros((n_storage, n_storage))
    
    return allocation_matrix, result

def write_allocation_excel(primary_storage_gdf, travel_time_matrix, distance_matrix, allocation_matrix, output_path):
    """
    Write storage allocation results to Excel workbook with three sheets.
    """
    ps = primary_storage_gdf.copy()
    storage_ids = ps.index.tolist()
    storage_labels = [f"{i}" for i in storage_ids]
    n_storage = len(ps)
    
    wb = Workbook()
    wb.remove(wb.active)
    
    # Sheet 1: Travel Time (minutes)
    ws1 = wb.create_sheet("storage_allocation_time")
    ws1.append(["Storage ID"] + storage_labels)
    for i, label_i in enumerate(storage_labels):
        row_data = [label_i] + travel_time_matrix[i, :].tolist()
        ws1.append(row_data)
    
    # Sheet 2: Distance (km)
    ws2 = wb.create_sheet("storage_allocation_distance")
    ws2.append(["Storage ID"] + storage_labels)
    for i, label_i in enumerate(storage_labels):
        row_data = [label_i] + distance_matrix[i, :].tolist()
        ws2.append(row_data)
    
    # Sheet 3: Allocation (percentages)
    ws3 = wb.create_sheet("storage_allocation_percentage")
    ws3.append(["Storage ID"] + storage_labels)
    for i, label_i in enumerate(storage_labels):
        row_data = [label_i] + allocation_matrix[i, :].tolist()
        ws3.append(row_data)
    
    wb.save(output_path)
    print(f"✓ Saved: {output_path}")
    
    return Path(output_path)

def allocation_process(primary_storage_gdf, friction_path, output_excel_path):
    """
    Orchestrate complete storage allocation optimization workflow.
    Builds routing matrices, solves LP, validates results, and exports to Excel.
    """
    ps = primary_storage_gdf.copy()
    n_storage = len(ps)
    
    print(f"Primary storage facilities: {n_storage}")
    print(f"Supply: {ps['percentage_supply'].sum():.4f}")
    print(f"Demand: {ps['percentage_demand'].sum():.4f}\n")
    print()
    
    # Step 1: Build routing matrices
    travel_time_matrix, distance_matrix, crs_obj, storage_ids = build_storage_routing_matrices(ps, friction_path)
    print()
    
    # Step 2: Optimize allocation
    allocation_matrix, opt_result = optimize_storage_allocation(ps, travel_time_matrix)
    print()
    
    # Step 3: Validate
    total_allocated = allocation_matrix.sum()
    total_supply = ps['percentage_supply'].sum()
    total_demand = ps['percentage_demand'].sum()
    
    print(f"Validation:")
    print(f"  Total allocation: {total_allocated:.6f}")
    print(f"  Total supply: {total_supply:.6f}")
    print(f"  Total demand: {total_demand:.6f}\n")
    
    if abs(total_allocated - total_supply) > 1e-6 or abs(total_allocated - total_demand) > 1e-6:
        print(f"  ⚠ Warning: allocation does not balance!\n")
    else:
        print(f"  ✓ Allocation balanced\n")
    print()
    
    # Step 4: Export
    write_allocation_excel(ps, travel_time_matrix, distance_matrix, allocation_matrix, output_excel_path)
    
    return {
        'allocation_matrix': allocation_matrix,
        'travel_time_matrix': travel_time_matrix,
        'distance_matrix': distance_matrix,
        'optimization_result': opt_result,
        'n_storage': n_storage,
        'total_supply': total_supply,
        'total_demand': total_demand,
    }

# Mattia part

import numpy as np
import pandas as pd
import geopandas as gpd
from rasterio.warp import Resampling
from rasterio.features import geometry_mask
import function_tools as tool

from parameters import (
    WALK_THRESHOLD_MIN, CAR_THRESHOLD_MIN, URBAN_SHARE, RURAL_SHARE,
    URBAN_THRESHOLD, NODATA_FLOAT, NODATA_INT
)

def calibrate_demand_and_allocate_clients(
    pixel_pref_path, pop_path, urban_path, boundary_path,
    resell_gdf, filling_gdf, output_lpg_use_path
):
    """
    Calibrates LPG demand based on urban/rural targets, generates an LPG use raster,
    allocates clients to resellers, and aggregates total demand at filling points.
    Returns: updated_resell_gdf, updated_filling_gdf
    """
    # 1. Load and Prepare Spatial Data
    pixel_pref = tool._read_pixel_preference_raster(pixel_pref_path)
    pop, pop_profile, _ = tool._read_raster(pop_path)
    transform, crs = pop_profile["transform"], pop_profile["crs"]
    height, width = pop.shape

    urban = tool.reproject_to_reference(urban_path, pop_profile, Resampling.nearest)
    urban = np.where(np.isfinite(urban), urban, np.nan).astype(np.float32)

    boundary = gpd.read_file(boundary_path)
    if boundary.crs != crs:
        boundary = boundary.to_crs(crs)
        
    mask_nigeria = geometry_mask(
        [boundary.geometry.union_all()], transform=transform, invert=True, out_shape=(height, width)
    )
    pop = np.where(np.isfinite(pop), np.maximum(pop, 0.0), np.nan).astype(np.float32)

    # 2. Extract Valid Pixels & Calculate Client Shares
    valid_pref = np.isfinite(pixel_pref["walk_share"]) & np.isfinite(pixel_pref["car_share"])
    valid_pref &= ((pixel_pref["best_reseller_id_walk"] >= 0) | (pixel_pref["best_reseller_id_car"] >= 0))
    
    rows, cols = np.where(valid_pref)
    rows, cols = rows.astype(np.int64), cols.astype(np.int64)

    if rows.size == 0:
        raise RuntimeError("No valid preference pixels found in multilayer raster.")

    pixel_pop = pop[rows, cols].astype(np.float64)
    valid_pop = np.isfinite(pixel_pop) & (pixel_pop > 0)
    rows, cols, pixel_pop = rows[valid_pop], cols[valid_pop], pixel_pop[valid_pop]

    walk_share = np.clip(pixel_pref["walk_share"][rows, cols].astype(np.float64), 0.0, 1.0)
    car_share = np.clip(pixel_pref["car_share"][rows, cols].astype(np.float64), 0.0, 1.0)
    share_sum = walk_share + car_share

    valid_share = share_sum > 0
    if not np.any(valid_share):
        raise RuntimeError("Invalid walk/car shares: all pixels have zero total share.")
    if not np.all(valid_share):
        dropped = int(np.count_nonzero(~valid_share))
        print(f"Warning: dropping {dropped:,} pixels with zero total share.")
        rows = rows[valid_share]
        cols = cols[valid_share]
        pixel_pop = pixel_pop[valid_share]
        walk_share = walk_share[valid_share]
        car_share = car_share[valid_share]
        share_sum = share_sum[valid_share]

    walk_share /= share_sum
    car_share /= share_sum

    walk_time = pixel_pref["best_time_walk_min"][rows, cols].astype(np.float64)
    car_time = pixel_pref["best_time_car_min"][rows, cols].astype(np.float64)
    
    walk_reseller_id = np.where(np.isfinite(pixel_pref["best_reseller_id_walk"][rows, cols]), pixel_pref["best_reseller_id_walk"][rows, cols], NODATA_INT).astype(np.int64)
    car_reseller_id = np.where(np.isfinite(pixel_pref["best_reseller_id_car"][rows, cols]), pixel_pref["best_reseller_id_car"][rows, cols], NODATA_INT).astype(np.int64)

    walk_eligible = np.isfinite(walk_time) & (walk_time <= WALK_THRESHOLD_MIN)
    car_eligible = np.isfinite(car_time) & (car_time <= CAR_THRESHOLD_MIN)

    urban_pixel = urban[rows, cols]
    urban_flag = np.where(np.isfinite(urban_pixel), urban_pixel >= URBAN_THRESHOLD, False)
    rural_flag = ~urban_flag

    urban_pop_total = float(np.nansum(pixel_pop[urban_flag]))
    rural_pop_total = float(np.nansum(pixel_pop[rural_flag]))

    if urban_pop_total <= 0 or rural_pop_total <= 0:
        raise RuntimeError("Urban or rural population total is zero. Check urban/pop alignment.")

    urban_eligible_pop = float(np.nansum(pixel_pop[urban_flag] * (walk_share[urban_flag] * walk_eligible[urban_flag] + car_share[urban_flag] * car_eligible[urban_flag])))
    rural_eligible_pop = float(np.nansum(pixel_pop[rural_flag] * (walk_share[rural_flag] * walk_eligible[rural_flag] + car_share[rural_flag] * car_eligible[rural_flag])))

    urban_factor = (URBAN_SHARE * urban_pop_total) / urban_eligible_pop
    rural_factor = (RURAL_SHARE * rural_pop_total) / rural_eligible_pop

    eligible_weight = walk_share * walk_eligible.astype(np.float64) + car_share * car_eligible.astype(np.float64)
    region_factor = np.where(urban_flag, urban_factor, rural_factor)

    lpg_use_share_pixel = np.clip(region_factor * eligible_weight, 0.0, 1.0)
    clients_walk_pixel = pixel_pop * walk_share * walk_eligible.astype(np.float64) * region_factor
    clients_car_pixel = pixel_pop * car_share * car_eligible.astype(np.float64) * region_factor
    clients_max_ideal_walk_pixel = pixel_pop * walk_share
    clients_max_ideal_car_pixel = pixel_pop * car_share

    # 3. Save LPG Use Raster (Still writes to disk as requested)
    lpg_use_share_raster = np.zeros((height, width), dtype=np.float32)
    lpg_use_share_raster[~mask_nigeria] = NODATA_FLOAT
    lpg_use_share_raster[rows, cols] = lpg_use_share_pixel.astype(np.float32)
    out_profile = pop_profile.copy()
    out_profile.update(dtype="float32", count=1, nodata=NODATA_FLOAT)
    with rasterio.open(output_lpg_use_path, "w", **out_profile) as dst:
        dst.write(lpg_use_share_raster.astype(np.float32), 1)

    # 4. Process Reseller Clients (In-Memory)
    reseller_df = resell_gdf.copy()
    if reseller_df.crs != crs:
        reseller_df = reseller_df.to_crs(crs)
        
    reseller_df = reseller_df[reseller_df.geometry.notna() & (reseller_df.geometry.geom_type == "Point")].copy()
    reseller_df["reseller_id"] = tool._read_reseller_ids(reseller_df).astype(np.int64)

    walk_clients = (
        pd.DataFrame({"reseller_id": walk_reseller_id, "clients_walk": clients_walk_pixel, "clients_max_ideal_walk": clients_max_ideal_walk_pixel})
        .loc[lambda d: d["reseller_id"] >= 0].groupby("reseller_id", as_index=False).sum()
    )
    car_clients = (
        pd.DataFrame({"reseller_id": car_reseller_id, "clients_car": clients_car_pixel, "clients_max_ideal_car": clients_max_ideal_car_pixel})
        .loc[lambda d: d["reseller_id"] >= 0].groupby("reseller_id", as_index=False).sum()
    )

    clients_agg = walk_clients.merge(car_clients, on="reseller_id", how="outer").fillna(0.0)
    clients_agg["clients"] = clients_agg["clients_walk"] + clients_agg["clients_car"]
    clients_agg["clients_max_ideal"] = clients_agg["clients_max_ideal_walk"] + clients_agg["clients_max_ideal_car"]

    client_cols = ["clients_walk", "clients_car", "clients", "clients_max_ideal_walk", "clients_max_ideal_car", "clients_max_ideal"]
    reseller_out = reseller_df.merge(clients_agg, on="reseller_id", how="left")
    for col in client_cols:
        reseller_out[col] = reseller_out[col].fillna(0.0).astype(np.float64)

    reseller_out = gpd.GeoDataFrame(reseller_out, geometry="geometry", crs=crs)

    # 5. Build filling_point_clients (In-Memory)
    filling_44 = filling_gdf.copy()
    filling_44["id_res&fil"] = pd.to_numeric(filling_44["id_res&fil"], errors="coerce")
    filling_44 = filling_44[filling_44["id_res&fil"].notna() & (filling_44["id_res&fil"] > 0)].copy()
    filling_44["id_res&fil"] = filling_44["id_res&fil"].astype("int64")

    filling_base_cols = [c for c in ["id_res&fil", "place_id", "lat", "lon", "geometry"] if c in filling_44.columns]
    filling_base = filling_44[filling_base_cols].copy()
    filling_id_set = set(filling_base["id_res&fil"].tolist())

    sell_points = reseller_out.copy()
    sell_points["id_res&fil"] = pd.to_numeric(sell_points["id_res&fil"], errors="coerce")
    sell_points = sell_points[sell_points["id_res&fil"].notna() & (sell_points["id_res&fil"] > 0)].copy()
    sell_points["id_res&fil"] = sell_points["id_res&fil"].astype("int64")
    
    for c in client_cols:
        sell_points[c] = pd.to_numeric(sell_points[c], errors="coerce").fillna(0.0).astype(float)

    marker_col = next((c for c in ["id_filling_only", "id_fillingonly", "id_fillingOnly"] if c in sell_points.columns), None)
    local_cols = ["id_res&fil"] + client_cols + [c for c in ["place_id", "lat", "lon"] if c in sell_points.columns and c not in ["id_res&fil"] + client_cols]
    if marker_col and marker_col not in local_cols:
        local_cols.append(marker_col)

    local_metrics = sell_points[local_cols].drop_duplicates(subset=["id_res&fil"], keep="first")
    out = filling_base.merge(local_metrics, on="id_res&fil", how="left", suffixes=("", "_sp"))
    for c in client_cols:
        out[c] = pd.to_numeric(out[c], errors="coerce").fillna(0.0).astype(float)

    if marker_col is None:
        out["id_filling_only"] = pd.NA
        marker_col = "id_filling_only"

    resell_44 = resell_gdf.copy()
    res_assign = resell_44[["id_res&fil", "nearest_filling_points_id"]].copy()
    res_assign = res_assign.apply(pd.to_numeric, errors="coerce").dropna()
    res_assign = res_assign[(res_assign > 0).all(axis=1)].astype("int64")
    res_assign = res_assign[res_assign["nearest_filling_points_id"].isin(filling_id_set)].copy()

    point_clients = sell_points[["id_res&fil", "clients", "clients_max_ideal"]].copy()
    assigned = res_assign.merge(point_clients, on="id_res&fil", how="left").fillna(0.0)

    agg = (
        assigned.groupby("nearest_filling_points_id", as_index=False)[["clients", "clients_max_ideal"]]
        .sum().rename(columns={"nearest_filling_points_id": "id_res&fil", "clients": "assigned_fil_clients", "clients_max_ideal": "assigned_fil_max_ideal_clients"})
    )

    out = out.merge(agg, on="id_res&fil", how="left").fillna(0.0)
    out["total_fil_clients"] = out["clients"] + out["assigned_fil_clients"]
    out["total_max_ideal_clients"] = out["clients_max_ideal"] + out["assigned_fil_max_ideal_clients"]

    keep_cols = [c for c in ["id_res&fil", marker_col, "place_id", "lat", "lon"] + client_cols + ["total_fil_clients", "total_max_ideal_clients", "geometry"] if c in out.columns]
    out = gpd.GeoDataFrame(out[keep_cols].copy(), geometry="geometry", crs=filling_44.crs)

    return reseller_out, out

def connected_components_8(mask: np.ndarray) -> np.ndarray:
    """
    Label 8‑connected components of True cells.
    Returns an integer array (0 = background).
    Falls back to pure Python if scipy is not available.
    """
    # Try to use scipy first (much faster)
    try:
        from scipy import ndimage
        structure = np.ones((3, 3), dtype=np.uint8)
        labels, _ = ndimage.label(mask, structure=structure)
        return labels.astype(np.int32)
    except ImportError:
        # Fallback pure Python implementation
        h, w = mask.shape
        labels = np.zeros((h, w), dtype=np.int32)
        current = 0
        stack = []
        for r in range(h):
            for c in range(w):
                if not mask[r, c] or labels[r, c] != 0:
                    continue
                current += 1
                labels[r, c] = current
                stack.append((r, c))
                while stack:
                    rr, cc = stack.pop()
                    for dr, dc, _ in NEIGHBORS:
                        nr, nc = rr + dr, cc + dc
                        if nr < 0 or nr >= h or nc < 0 or nc >= w:
                            continue
                        if not mask[nr, nc] or labels[nr, nc] != 0:
                            continue
                        labels[nr, nc] = current
                        stack.append((nr, nc))
        return labels

def load_reseller_seeds(
    gpkg_path: Path,
    layer: str,
    id_col: str,
    attr_col: str,
    ref_profile: Dict,
) -> List[Dict]:
    """
    Load reseller points, reproject to reference CRS, and keep only those
    falling inside the raster extent. If multiple points map to the same pixel,
    keep the most attractive one.
    Returns a list of dicts with keys: reseller_id, attractiveness, row, col.
    """
    gdf = gpd.read_file(gpkg_path, layer=layer)
    if gdf.empty:
        raise ValueError("Reseller layer is empty.")
    ref_crs = ref_profile["crs"]
    if gdf.crs is None:
        raise ValueError("Reseller GeoPackage has no CRS.")
    if gdf.crs != ref_crs:
        gdf = gdf.to_crs(ref_crs)
    # Clean IDs and attractiveness
    ids = pd.to_numeric(gdf[id_col], errors="coerce")
    attrs = pd.to_numeric(gdf[attr_col], errors="coerce")
    attrs = attrs.fillna(ATTR_MIN).clip(lower=ATTR_MIN)
    xs = gdf.geometry.x.to_numpy()
    ys = gdf.geometry.y.to_numpy()
    rows, cols = rowcol(ref_profile["transform"], xs, ys)
    h, w = ref_profile["height"], ref_profile["width"]
    seeds_dict = {}
    for rid, attr, r, c in zip(ids, attrs, rows, cols):
        if not np.isfinite(rid):
            continue
        rr = int(r)
        cc = int(c)
        if rr < 0 or rr >= h or cc < 0 or cc >= w:
            continue
        key = (rr, cc)
        candidate = {
            "reseller_id": float(rid),
            "attractiveness": float(attr),
            "row": rr,
            "col": cc,
        }
        prev = seeds_dict.get(key)
        if prev is None or candidate["attractiveness"] > prev["attractiveness"]:
            seeds_dict[key] = candidate
    seeds = list(seeds_dict.values())
    if not seeds:
        raise ValueError("No reseller seed mapped onto the reference grid.")
    return seeds

def run_multisource_dijkstra(
    mode_name: str,
    friction: np.ndarray,
    traversable: np.ndarray,
    target_mask: np.ndarray,
    labels: np.ndarray,
    seeds: List[Dict],
    pixel_size_m: float,
    log_every: int | None = None,
) -> Tuple[np.ndarray, np.ndarray, np.ndarray, int, float]:
    """
    Multi‑source Dijkstra con Huff‑style scoring (β=2).
    Returns:
        best_seed_idx : 2D array of seed indices (or -1)
        best_time_min : 2D array of travel times in minutes (or inf)
        best_dist_km  : 2D array of path distance in km (or inf)
        assigned      : number of target pixels assigned
        elapsed_sec   : solver wall time
    """
    if log_every is None:
        log_every = LOG_EVERY
    h, w = friction.shape
    n_seeds = len(seeds)
    # Precompute source data
    seed_rows = np.array([s["row"] for s in seeds], dtype=np.int32)
    seed_cols = np.array([s["col"] for s in seeds], dtype=np.int32)
    sqrt_attr = np.array(
        [math.sqrt(max(s["attractiveness"], ATTR_MIN)) for s in seeds],
        dtype=np.float64,
    )
    # Determine which connected components contain at least one seed
    seed_labels = labels[seed_rows, seed_cols]
    max_label = labels.max()
    has_seed = np.zeros(max_label + 1, dtype=bool)
    has_seed[seed_labels[seed_labels > 0]] = True
    # State arrays
    best_adj = np.full((h, w), np.inf, dtype=np.float64)   # adjusted time
    best_time = np.full((h, w), np.inf, dtype=np.float64)  # raw time (minutes)
    best_dist = np.full((h, w), np.inf, dtype=np.float64)  # distance (km)
    best_seed = np.full((h, w), -1, dtype=np.int32)
    finalized = np.zeros((h, w), dtype=bool)
    # Priority queue: (adj_time, row, col, seed_idx)
    heap = []
    for si in range(n_seeds):
        r = seed_rows[si]
        c = seed_cols[si]
        if not traversable[r, c]:
            continue
        # At source pixel, tie‑break in favour of higher attractiveness
        if (0.0 < best_adj[r, c]) or (
            best_adj[r, c] == 0.0
            and best_seed[r, c] >= 0
            and sqrt_attr[si] > sqrt_attr[best_seed[r, c]]
        ):
            best_adj[r, c] = 0.0
            best_time[r, c] = 0.0
            best_dist[r, c] = 0.0
            best_seed[r, c] = si
            heapq.heappush(heap, (0.0, r, c, si))
    total_targets = int(np.count_nonzero(target_mask & traversable))
    assigned = 0
    t_start = time.perf_counter()
    last_log = 0
    while heap and assigned < total_targets:
        adj, r, c, si = heapq.heappop(heap)
        if finalized[r, c]:
            continue
        if adj > best_adj[r, c]:
            continue
        finalized[r, c] = True
        if target_mask[r, c]:
            assigned += 1
            if assigned - last_log >= log_every:
                tool._print_progress(
                    mode_name,
                    assigned,
                    total_targets,
                    time.perf_counter() - t_start,
                )
                last_log = assigned
        inv_scale = 1.0 / sqrt_attr[si]
        t_curr = best_time[r, c]
        d_curr = best_dist[r, c]
        for dr, dc, dist_factor in NEIGHBORS:
            nr = r + dr
            nc = c + dc
            if nr < 0 or nr >= h or nc < 0 or nc >= w:
                continue
            if finalized[nr, nc] or not traversable[nr, nc]:
                continue
            # Edge cost: average friction × distance (metres) → minutes
            edge_time = float(
                0.5 * (friction[r, c] + friction[nr, nc])
                * (pixel_size_m * dist_factor)
            )
            t_new = t_curr + edge_time
            adj_new = t_new * inv_scale
            # Distance in km (geometric, independent of friction)
            edge_km = (pixel_size_m * dist_factor) / 1000.0
            d_new = d_curr + edge_km
            # Lexicographic tie‑break: lower adjusted time, then lower raw time
            if (adj_new + 1e-12) < best_adj[nr, nc] or (
                abs(adj_new - best_adj[nr, nc]) <= 1e-12
                and t_new < best_time[nr, nc]
            ):
                best_adj[nr, nc] = adj_new
                best_time[nr, nc] = t_new
                best_dist[nr, nc] = d_new
                best_seed[nr, nc] = si
                heapq.heappush(heap, (adj_new, nr, nc, si))
    # -------------------------------------------------------------------------
    # Component‑aware fallback
    # -------------------------------------------------------------------------
    unassigned = target_mask & traversable & (best_seed < 0)
    eligible = unassigned & has_seed[labels]
    fallback_hits = 0
    if np.any(eligible):
        unique_lbls = np.unique(labels[eligible])
        for lbl in unique_lbls:
            if lbl <= 0:
                continue
            comp_mask = labels == lbl
            comp_targets = eligible & comp_mask
            if not np.any(comp_targets):
                continue
            comp_seed_idx = np.where(seed_labels == lbl)[0]
            if comp_seed_idx.size == 0:
                continue
            fallback_hits += int(np.count_nonzero(comp_targets))
            # Reset state inside the component
            best_adj_comp = np.full((h, w), np.inf, dtype=np.float64)
            best_time_comp = np.full((h, w), np.inf, dtype=np.float64)
            best_dist_comp = np.full((h, w), np.inf, dtype=np.float64)
            best_seed_comp = np.full((h, w), -1, dtype=np.int32)
            finalized_comp = np.zeros((h, w), dtype=bool)
            heap_comp = []
            for si in comp_seed_idx:
                r0 = seed_rows[si]
                c0 = seed_cols[si]
                if not comp_mask[r0, c0] or not traversable[r0, c0]:
                    continue
                best_adj_comp[r0, c0] = 0.0
                best_time_comp[r0, c0] = 0.0
                best_dist_comp[r0, c0] = 0.0
                best_seed_comp[r0, c0] = si
                heapq.heappush(heap_comp, (0.0, r0, c0, si))
            while heap_comp:
                adj, r, c, si = heapq.heappop(heap_comp)
                if finalized_comp[r, c]:
                    continue
                if adj > best_adj_comp[r, c]:
                    continue
                if not comp_mask[r, c]:
                    continue
                finalized_comp[r, c] = True
                inv_scale = 1.0 / sqrt_attr[si]
                t_curr = best_time_comp[r, c]
                d_curr = best_dist_comp[r, c]
                for dr, dc, dist_factor in NEIGHBORS:
                    nr = r + dr
                    nc = c + dc
                    if nr < 0 or nr >= h or nc < 0 or nc >= w:
                        continue
                    if (
                        finalized_comp[nr, nc]
                        or not traversable[nr, nc]
                        or not comp_mask[nr, nc]
                    ):
                        continue
                    edge_time = float(
                        0.5 * (friction[r, c] + friction[nr, nc])
                        * (pixel_size_m * dist_factor)
                    )
                    t_new = t_curr + edge_time
                    adj_new = t_new * inv_scale
                    edge_km = (pixel_size_m * dist_factor) / 1000.0
                    d_new = d_curr + edge_km
                    if (adj_new + 1e-12) < best_adj_comp[nr, nc] or (
                        abs(adj_new - best_adj_comp[nr, nc]) <= 1e-12
                        and t_new < best_time_comp[nr, nc]
                    ):
                        best_adj_comp[nr, nc] = adj_new
                        best_time_comp[nr, nc] = t_new
                        best_dist_comp[nr, nc] = d_new
                        best_seed_comp[nr, nc] = si
                        heapq.heappush(heap_comp, (adj_new, nr, nc, si))
            # Merge fallback results into global arrays
            update = comp_targets & (best_seed_comp >= 0)
            best_seed[update] = best_seed_comp[update]
            best_time[update] = best_time_comp[update]
            best_dist[update] = best_dist_comp[update]
    elapsed = time.perf_counter() - t_start
    assigned_final = int(np.count_nonzero(target_mask & traversable & (best_seed >= 0)))
    tool._print_progress(mode_name, assigned_final, total_targets, elapsed)
    return best_seed, best_time, best_dist, assigned_final, elapsed

def seeds_to_raster(
    best_seed_idx: np.ndarray,
    best_time_min: np.ndarray,
    best_dist_km: np.ndarray,
    seeds: List[Dict],
    inhabited_mask: np.ndarray,
) -> Tuple[np.ndarray, np.ndarray, np.ndarray]:
    """Convert seed indices, times and distances to final rasters."""
    h, w = best_seed_idx.shape
    id_raster = np.full((h, w), NODATA, dtype=np.float32)
    time_raster = np.full((h, w), NODATA, dtype=np.float32)
    dist_raster = np.full((h, w), NODATA, dtype=np.float32)
    valid = inhabited_mask & (best_seed_idx >= 0) & np.isfinite(best_time_min)
    if np.any(valid):
        seed_ids = np.array([s["reseller_id"] for s in seeds], dtype=np.float64)
        idx = best_seed_idx[valid]
        id_raster[valid] = seed_ids[idx].astype(np.float32)
        time_raster[valid] = best_time_min[valid].astype(np.float32)
        dist_raster[valid] = best_dist_km[valid].astype(np.float32)
    return id_raster, time_raster, dist_raster

def recompute_exact_metrics(
    mode_name: str,
    friction: np.ndarray,
    traversable: np.ndarray,
    id_raster: np.ndarray,                # Phase‑1 assignment (IDs, NODATA elsewhere)
    phase1_time_raster: np.ndarray,       # approximate times from Phase 1 (for limit)
    seeds: List[Dict],
    pixel_size_m: float,
    time_limit_factor: float | None = None,
    max_distance_km: float = EXACT_MAX_DISTANCE_KM,
    log_every: int = 500,
) -> Tuple[np.ndarray, np.ndarray]:
    """
    Recompute exact travel times and distances for pixels assigned to each reseller.
    For every seed, run a single‑source Dijkstra (no finalisation) and record
    the true metrics only for pixels whose Phase‑1 assignment matches the seed.
    Early termination:
        - Stop when all assigned pixels of the current seed have been reached.
        - Alternatively, stop when the time exceeds `max_phase1_time * time_limit_factor`.
          If this second condition triggers, a warning is printed.
    Returns:
        exact_time_raster : 2D array with exact travel times (minutes, NODATA for unassigned)
        exact_dist_raster : 2D array with exact distances (km, NODATA for unassigned)
    """
    seeds_with_dist_warning = set()
    seeds_with_time_warning = set()
    total_target_pixels = 0
    total_skipped_pixels = 0

    if time_limit_factor is None:
        time_limit_factor = TIME_LIMIT_FACTOR
    h, w = friction.shape
    exact_time = np.full((h, w), NODATA, dtype=np.float32)
    exact_dist = np.full((h, w), NODATA, dtype=np.float32)
    # Build mapping from seed ID to its index
    seed_id_to_idx = {}
    for si, s in enumerate(seeds):
        sid = s["reseller_id"]
        if sid not in seed_id_to_idx:
            seed_id_to_idx[sid] = si
    assigned_mask = (id_raster != NODATA) & traversable
    unique_ids = np.unique(id_raster[assigned_mask])
    n_seeds_total = len(unique_ids)
    n_processed = 0
    t_start = time.perf_counter()
    last_log = 0
    print(f"[{mode_name}-exact] Processing {n_seeds_total:,} resellers...")
    for sid in unique_ids:
        si = seed_id_to_idx.get(sid)
        if si is None:
            continue
        seed = seeds[si]
        sr, sc = seed["row"], seed["col"]
        if not traversable[sr, sc]:
            continue
        target_mask = assigned_mask & (id_raster == sid)
        target_coords = np.argwhere(target_mask)  # list of (r,c)
        target_count = len(target_coords)
        total_target_pixels += target_count
        if target_count == 0:
            continue
        # Determine maximum allowed time for early termination
        phase1_times = phase1_time_raster[target_mask]
        max_phase1_time = np.max(phase1_times) if len(phase1_times) > 0 else 0.0
        time_limit = max_phase1_time * time_limit_factor if max_phase1_time > 0 else np.inf
        # Dijkstra state
        dist_time = np.full((h, w), np.inf, dtype=np.float64)
        dist_dist = np.full((h, w), np.inf, dtype=np.float64)
        visited = np.zeros((h, w), dtype=bool)
        dist_time[sr, sc] = 0.0
        dist_dist[sr, sc] = 0.0
        heap = [(0.0, sr, sc)]
        reached_targets = 0
        target_set = set((r, c) for r, c in target_coords)
        while heap:
            t, r, c = heapq.heappop(heap)
            if visited[r, c]:
                continue
            visited[r, c] = True
            # If time limit exceeded (and we haven't reached all targets), stop early
            if t > time_limit and reached_targets < target_count:
                if sid not in seeds_with_time_warning:
                    seeds_with_time_warning.add(sid)
                break
            # If this cell is a target for this seed, record exact metrics
            if (r, c) in target_set:
                exact_time[r, c] = t
                exact_dist[r, c] = dist_dist[r, c]
                reached_targets += 1
                if reached_targets == target_count:
                    break  # all done
            t_curr = t
            d_curr = dist_dist[r, c]
            # Distance‑based early termination
            if d_curr > max_distance_km:
                if sid not in seeds_with_dist_warning:
                    seeds_with_dist_warning.add(sid)
                continue
            for dr, dc, dist_factor in NEIGHBORS:
                nr = r + dr
                nc = c + dc
                if nr < 0 or nr >= h or nc < 0 or nc >= w:
                    continue
                if visited[nr, nc] or not traversable[nr, nc]:
                    continue
                edge_time = float(
                    0.5 * (friction[r, c] + friction[nr, nc])
                    * (pixel_size_m * dist_factor)
                )
                t_new = t_curr + edge_time
                edge_km = (pixel_size_m * dist_factor) / 1000.0
                d_new = d_curr + edge_km
                if t_new < dist_time[nr, nc] - 1e-12:
                    dist_time[nr, nc] = t_new
                    dist_dist[nr, nc] = d_new
                    heapq.heappush(heap, (t_new, nr, nc))
        # End while heap
        total_skipped_pixels += (target_count - reached_targets)
        n_processed += 1
        if n_processed - last_log >= log_every:
            elapsed = time.perf_counter() - t_start
            pct = 100.0 * n_processed / n_seeds_total
            rate = tool._safe_ratio(n_processed, elapsed)
            remaining = n_seeds_total - n_processed
            eta = tool._safe_ratio(remaining, max(rate, 1e-12))
            print(
                f"[{mode_name}-exact] {n_processed:,}/{n_seeds_total:,} resellers done "
                f"({pct:.1f}%) | {rate:.1f} res/s | ETA {eta/60:.1f} min"
            )
            last_log = n_processed
    # End loop over seeds
    elapsed = time.perf_counter() - t_start
    # Print warning summary
    if seeds_with_dist_warning:
        print(f"  Warning: {len(seeds_with_dist_warning)}/{n_seeds_total} seeds ({100*len(seeds_with_dist_warning)/n_seeds_total:.1f}%) exceeded distance limit {max_distance_km} km")
    if seeds_with_time_warning:
        print(f"  Warning: {len(seeds_with_time_warning)}/{n_seeds_total} seeds ({100*len(seeds_with_time_warning)/n_seeds_total:.1f}%) exceeded time limit (factor {time_limit_factor})")
    if total_target_pixels > 0:
        skip_pct = 100.0 * total_skipped_pixels / total_target_pixels
        print(f"  Skipped pixels: {total_skipped_pixels}/{total_target_pixels} ({skip_pct:.1f}%) due to distance or time limits.")
    print(f"[{mode_name}-exact] All resellers processed. Assigned pixels recomputed.")
    return exact_time, exact_dist


import numpy as np
import geopandas as gpd
import rasterio
from rasterio.transform import rowcol
from typing import Optional, List, Dict

def assign_best(
    origins: gpd.GeoDataFrame,
    origin_id_col: str,
    destinations: Optional[gpd.GeoDataFrame] = None,
    layer1_name: str = "origin",
    friction_path: Optional[str] = None,
    surface_data: Optional[tuple] = None,
    huff: bool = False,
    attractiveness_col: Optional[str] = None,
    output_raster_path: Optional[str] = None,
    pixel_size_m: Optional[float] = None,
    pop_mask: Optional[str] = None,
    avg_speed_band: bool = False
) -> Optional[gpd.GeoDataFrame]:
    """
    Uses the existing multisource Dijkstra (which internally applies sqrt(attractiveness)
    scoring when attractiveness differs from 1.0). Exact recomputation is performed
    only when huff=True.
    """
    # 1. Establish Master Reference Profile (forces EPSG:3857 if pop_mask is 3857)
    ref_profile = None
    if pop_mask is not None:
        _, ref_profile, _ = tool._read_raster_base(pop_mask)

    # 2. Build / retrieve cost surface aligned to reference
    if surface_data is None:
        if friction_path is None:
            raise ValueError("Either surface_data or friction_path must be provided")
        surface_data = build_surface(friction_path, ref_profile=ref_profile)
        
    cost_layer, transform, geod, crs_obj = surface_data

    # 3. Dynamically calculate pixel size from transform if not provided
    if pixel_size_m is None:
        pixel_size_m = abs(transform[0])

    if huff:
        if attractiveness_col is None:
            raise ValueError("attractiveness_col is required when huff=True")
        if attractiveness_col not in origins.columns:
            raise ValueError(f"Column '{attractiveness_col}' not in origins")

    if destinations is None and output_raster_path is None:
        raise ValueError("At least one of 'destinations' or 'output_raster_path' must be given")

    # 2. Reproject to raster CRS
    origins_proj = origins.to_crs(crs_obj)
    dests_proj = None
    if destinations is not None:
        dests_proj = destinations.to_crs(crs_obj)

    # 3. Traversable mask (valid friction cells)
    traversable = np.isfinite(cost_layer) & (cost_layer > 0.0)
    target_mask = np.zeros_like(traversable, dtype=bool)

    def _pixel_coords(gdf):
        rows, cols = [], []
        for geom in gdf.geometry:
            if geom.is_empty:
                rows.append(-1); cols.append(-1)
            else:
                r, c = rowcol(transform, geom.x, geom.y)
                rows.append(r); cols.append(c)
        return np.array(rows, dtype=np.int32), np.array(cols, dtype=np.int32)

    o_rows, o_cols = _pixel_coords(origins_proj)
    o_valid = traversable[o_rows, o_cols] & (o_rows >= 0)
    o_rows = o_rows[o_valid]; o_cols = o_cols[o_valid]
    o_ids = origins_proj[origin_id_col].iloc[o_valid].to_numpy()

    if huff:
        o_attr = origins_proj[attractiveness_col].iloc[o_valid].to_numpy(dtype=float)
    else:
        o_attr = np.ones(len(o_rows), dtype=float)

    d_rows = d_cols = None
    dest_indices_valid = np.array([], dtype=int)
    if dests_proj is not None:
        d_rows, d_cols = _pixel_coords(dests_proj)
        d_valid = traversable[d_rows, d_cols] & (d_rows >= 0)
        d_rows = d_rows[d_valid]; d_cols = d_cols[d_valid]
        dest_indices_valid = np.where(d_valid)[0]
        target_mask[d_rows, d_cols] = True

    if output_raster_path is not None and pop_mask is not None:
        pop_aligned = np.zeros_like(cost_layer, dtype=np.float32)
        with rasterio.open(pop_mask) as src:
            reproject(
                source=rasterio.band(src, 1),
                destination=pop_aligned,
                src_transform=src.transform,
                src_crs=src.crs,
                dst_transform=transform,
                dst_crs=crs_obj,
                resampling=Resampling.nearest
            )
        target_mask = target_mask | (traversable & (pop_aligned > 0))
    elif output_raster_path is not None:
        target_mask = traversable

    if len(o_rows) == 0:
        if dests_proj is not None:
            result = destinations.copy()
            result[f"nearest_{layer1_name}_id"] = None
            result[f"nearest_{layer1_name}_time"] = np.nan
            result[f"nearest_{layer1_name}_distance"] = np.nan
            return result
        return None

    # 4. Build seeds
    seeds: List[Dict] = []
    for i in range(len(o_rows)):
        seeds.append({
            "reseller_id": o_ids[i],
            "attractiveness": float(o_attr[i]),
            "row": int(o_rows[i]),
            "col": int(o_cols[i]),
        })

    # 5. Connected components (required by Dijkstra)
    labels = connected_components_8(traversable)

    mode_label = layer1_name or "assign"

    # 6. Run multisource Dijkstra (the original, with Huff scoring built in)
    best_seed_idx, best_time_approx, best_dist_approx, _, _ = run_multisource_dijkstra(
        mode_name=mode_label,
        friction=cost_layer,
        traversable=traversable,
        target_mask=target_mask,   
        labels=labels,
        seeds=seeds,
        pixel_size_m=pixel_size_m,
        log_every=250000,
    )

    # 7. Exact recomputation (only for Huff)
    if huff:
        id_raster = np.full(cost_layer.shape, -1, dtype=np.int32)
        for i, seed in enumerate(seeds):
            id_raster[best_seed_idx == i] = seed["reseller_id"]
        time_exact, dist_exact = recompute_exact_metrics(
            mode_name=mode_label,
            friction=cost_layer,
            traversable=traversable,
            id_raster=id_raster,
            phase1_time_raster=best_time_approx,
            seeds=seeds,
            pixel_size_m=pixel_size_m,
            log_every=500,
        )
        best_time = time_exact
        best_dist = dist_exact
    else:
        best_time = best_time_approx
        best_dist = best_dist_approx

    # 8. Vector output
    result_gdf = None
    if dests_proj is not None:
        result_gdf = destinations.copy()
        id_col = f"nearest_{layer1_name}_id"
        time_col = f"nearest_{layer1_name}_time"
        dist_col = f"nearest_{layer1_name}_distance"
        result_gdf[id_col] = None
        result_gdf[time_col] = np.nan
        result_gdf[dist_col] = np.nan

        for i, dest_idx in enumerate(dest_indices_valid):
            r, c = d_rows[i], d_cols[i]
            sidx = best_seed_idx[r, c]
            if sidx < 0:
                continue
            seed = seeds[sidx]
            t_val = best_time[r, c]
            d_val = best_dist[r, c]
            result_gdf.at[result_gdf.index[dest_idx], id_col] = seed["reseller_id"]
            if np.isfinite(t_val):
                result_gdf.at[result_gdf.index[dest_idx], time_col] = float(t_val)
            if np.isfinite(d_val):
                result_gdf.at[result_gdf.index[dest_idx], dist_col] = float(d_val)

    # 9. Raster output
    if output_raster_path is not None:
        out_id = np.full(cost_layer.shape, -9999, dtype=np.int32)
        out_time = np.full(cost_layer.shape, -9999.0, dtype=np.float32)
        out_dist = np.full(cost_layer.shape, -9999.0, dtype=np.float32)

        for i, seed in enumerate(seeds):
            mask = (best_seed_idx == i) & target_mask
            out_id[mask] = seed["reseller_id"]
            out_time[mask] = best_time[mask]
            out_dist[mask] = best_dist[mask]

        band_count = 4 if avg_speed_band else 3

        if friction_path:
            with rasterio.open(friction_path) as src:
                out_meta = src.meta.copy()
        else:
            out_meta = {
                "driver": "GTiff",
                "height": cost_layer.shape[0],
                "width": cost_layer.shape[1],
                "count": band_count,
                "dtype": "float32",
                "crs": crs_obj,
                "transform": transform,
                "nodata": -9999.0,
            }
            
        out_meta.update(driver="GTiff", count=band_count, nodata=-9999.0)
        
        band_names = [
            f"nearest_{layer1_name}_id",
            f"nearest_{layer1_name}_time",
            f"nearest_{layer1_name}_distance",
        ]

        if avg_speed_band:       # TODO rimuovi modalità avg_speed_band se non viene piu usata
            band_names.append(f"nearest_{layer1_name}_speed_kmh")
            out_speed = np.full(cost_layer.shape, -9999.0, dtype=np.float32)
            valid_mask = (out_time != -9999.0) & (out_time > 0)
            out_speed[valid_mask] = out_dist[valid_mask] / (out_time[valid_mask] / 60.0)
            out_speed[(out_time == 0) & (out_id != -9999)] = 0.0

        with rasterio.open(output_raster_path, "w", **out_meta) as dst:
            dst.write(out_id, 1)
            dst.write(out_time.astype(np.float32), 2)
            dst.write(out_dist.astype(np.float32), 3)
            if avg_speed_band:
                dst.write(out_speed.astype(np.float32), 4)
            dst.descriptions = band_names

    return result_gdf